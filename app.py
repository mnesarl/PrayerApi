# main.py
# Import sys Module
import sys
import os
import time
import datetime
import traceback
# os.system('python.exe -m pip install --upgrade pip')
# os.system("python3 -m pip install types-requests")
import sys
import pathlib

try:
    from dateutil import parser
except:
    os.system("pip install python-dateutil")
    os.execv(sys.executable, [sys.executable] + sys.argv)
try:
    import requests
except:
    os.system("pip install requests")
    os.execv(sys.executable, [sys.executable] + sys.argv)

import configparser
import openpyxl
from threading import *
from bs4 import BeautifulSoup

try:
    import pandas as pd  # pip install pandas
except:
    os.system('pip install pandas')
    os.execv(sys.executable, [sys.executable] + sys.argv)
# Import PyQt5 Module
from PyQt5 import QtCore, QtGui, QtWidgets, uic, QtNetwork
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtCore import QTimer, QDateTime, QUrl, pyqtSlot, QDir, Qt
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtNetwork import QNetworkAccessManager, QNetworkRequest
# from PyQt5.QtGui import (QFont,QFontDatabase)
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import *
import sqlite3
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog
from PyQt5.QtCore import QDir, Qt, QUrl
from PyQt5.QtWidgets import (QMainWindow, QWidget, QPushButton, QApplication,
                             QLabel, QFileDialog, QStyle, QVBoxLayout, QTableWidget, QTableWidgetItem)

# from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog
# Constante
SelectedRole = QtCore.Qt.UserRole + 1000


def enModeTest():
    global debug
    debug = False
    return True


debug = False
# class datetime(datetime.datetime):
#     def __divmod__(self, delta):
#         seconds = int((self - datetime.datetime.min).total_seconds())
#         remainder = datetime.timedelta(
#             seconds=seconds % delta.total_seconds(),
#             microseconds=self.microsecond,
#         )
#         quotient = self - remainder
#         return quotient, remainder
#
#     def __floordiv__(self, delta):
#         return divmod(self, delta)[0]
#
#     def __mod__(self, delta):
#         return divmod(self, delta)[1]


gdLaDateDuJour_ = datetime.datetime.now().strftime("%Y-%m-%d")
if debug:
    print("Date du jour: ", gdLaDateDuJour_)


def time_conversion(sec):
    sec_value = sec % (24 * 3600)
    hour_value = sec_value // 3600
    sec_value %= 3600
    min = sec_value // 60
    sec_value %= 60

    return hour_value, min


def convert(seconds):
    return time.strftime("%H:%M:%S", time.gmtime(n))


def isTimeFormat(input):
    try:
        time.strptime(input, '%H:%M:%S')
        return True
    except ValueError:
        return False


def createConnection():
    con = QSqlDatabase.addDatabase("QSQLITE")
    con.setDatabaseName("Parametres.sqlite")
    if not con.open():
        QMessageBox.critical(
            None,
            "App Name - Error!",
            "Database Error: %s" % con.lastError().databaseText(),
        )
        return False

    # Open the connection
    createTableQuery = QSqlQuery()
    createTableQuery.exec(
        """
        CREATE TABLE IF NOT EXISTS Parametres (
            id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
            Param_Nom VARCHAR(40) NOT NULL,
            Param_Valeur VARCHAR(50)
        )
        """
    )
    createTableQuery.clear()

    return True


def get_Horaire(Param_Date, Param_Horaire):
    rows = None
    try:
        conn = sqlite3.connect("Parametres.sqlite")
        c = conn.cursor()
        sqlquery = "SELECT %s FROM Horaires Where Jour_Date='%s'" % (Param_Horaire, Param_Date)
        # "SELECT * from test WHERE desig_pu= ? ", (str(searchrol),)
        result = c.execute(sqlquery)
        rows = result.fetchone()
        rows = str(rows[0])
        # print("%s : %s" % (Param_Horaire, rows))
    except Exception as e:
        print("Exception get horaire: ", e)
    return rows
    # SELECT Jour_Date, Jour, Mois_AR, Mois_FR,heure_fajr, heure_shourouq, heure_dhuhr,heure_asr,heure_maghrib,
    # heure_ishae


def get_param(Param_N):
    query = QSqlQuery()
    query.exec("SELECT Param_Valeur FROM Parametres Where Param_Nom='%s'" % Param_N)
    query.first()
    return query.value(0)


# Insert Param table
def Insert_Update_Param():
    try:
        #         con = QSqlDatabase.addDatabase("QSQLITE")
        #         con.setDatabaseName("Parametres.sqlite")
        #         con.open()
        # Creating a query for later execution using .prepare()
        insertDataQuery = QSqlQuery()
        insertDataQuery.prepare(
            """
            INSERT INTO Parametres (
                id,
                Param_Nom,
                Param_Valeur
            )
            VALUES (?,?,?)
            """
        )
        # Sample data
        config2 = configparser.ConfigParser()
        file = pathlib.Path("config/config.ini")
        if file.exists():
            config2.read('config/config.ini')
            gbMONVER = config2['DEFAULT']['version']
            gbMONURL1 = config2['DEFAULT']['Url_WS1']
            gbMONURL2 = config2['DEFAULT']['Url_WS2']
            gbtime_out = config2['DEFAULT']['timeout']
            gbcity = config2['DEFAULT']['city']
            gbResolutionW = config2['DEFAULT']['ResolutionW']
            gbResolutionH = config2['DEFAULT']['ResolutionH']
            gbbackground_image_url = config2['DEFAULT']['background_image_url']
            gbazan_mp3 = config2['DEFAULT']['azan_mp3']
            gbremining_Time = config2['DEFAULT']['remining_Time']
            gbnote_Journee = config2['DEFAULT']['note_Journee']
            gbnote_Externe = config2['DEFAULT']['note_Externe']
            gbmode = config2['DEFAULT']['mode']

        data = [
            ("1", "version", gbMONVER),
            ("2", "Url_WS1", gbMONURL1),
            ("3", "Url_WS2", gbMONURL2),
            ("4", "timeout", gbtime_out),
            ("5", "city", gbcity),
            ("6", "ResolutionW", gbResolutionW),
            ("7", "ResolutionH", gbResolutionH),
            ("8", "background_image_url", gbbackground_image_url),
            ("9", "azan_mp3", gbazan_mp3),
            ("10", "remining_Time", gbremining_Time),
            ("11", "mode", gbmode),
            ("12", "note_Journee", gbnote_Journee),
            ("13", "note_Externe", gbnote_Externe),
        ]
        # Use .addBindValue() to insert data
        for id, Param_Nom, Param_Valeur in data:
            insertDataQuery.addBindValue(id)
            insertDataQuery.addBindValue(Param_Nom)
            insertDataQuery.addBindValue(Param_Valeur)
            insertDataQuery.exec()
        insertDataQuery.finish()
        return True
    except Exception as e:
        print("Exception insertDataQuery: ", e)
        return False


def check_if_existe(file):
    file = pathlib.Path(file)
    if not file.exists():
        print("Unable to find %s", file)
        QMessageBox.critical(
            None,
            "App Name - Error!",
            "File Error: not found",
        )
        return False
    else:
        return True


class comboCompanies(QComboBox):
    def __init__(self, parent):
        super().__init__(parent)
        self.setStyleSheet('font-size: 25px')
        self.addItems(['Microsoft', 'Facebook', 'Apple', 'Google'])
        self.currentIndexChanged.connect(self.getComboValue)

    def getComboValue(self):
        print(self.currentText())
        # return self.currentText()


class Horaires_Window(QWidget):
    signal = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        loadUi("ui/AnotherWindow.ui", self)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("png/1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        self.setWindowTitle("Horaires Priere")
        self.browse.clicked.connect(self.openFile)  # type: ignore
        self.browse.clicked.connect(self.akcja2)
        self.tableWidget.cellClicked.connect(self.cellClick)
        self.supprimer.clicked.connect(self.supprimerTout)
        self.btn_supprimer.clicked.connect(self.deletecurrentrow)  # type: ignore
        self.btn_Valider.clicked.connect(self.valider)
        self.btn_Annuler.clicked.connect(self.loaddata)
        self.btn_Valider.hide()
        self.btn_Annuler.hide()
        if not enModeTest():
            self.supprimer.hide()
            self.btn_supprimer.hide()
        self.loaddata()

    def deletecurrentrow(self):
        self.tableWidget.removeRow(self.tableWidget.currentRow())

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Delete:
            row = self.currentRow()
            self.removeRow(row)
        else:
            super().keyPressEvent(event)

    def cellClick(self, row, col):
        self.row = row
        self.col = col
        print(row)
        print(col)

    def supprimerTout(self):
        connection = sqlite3.connect('Parametres.sqlite')
        cur = connection.cursor()
        if not cur:
            QMessageBox.critical(
                None,
                "App Name - Error!",
                "Database Error: %s" % con.lastError().databaseText(),
            )
        tablename = "Horaires"
        sqlstr = "DELETE FROM {} ".format(tablename)
        cur.execute(sqlstr)
        cur.close()
        connection.commit()
        connection.close()
        self.loaddata()

    def loaddata(self):
        connection = sqlite3.connect('Parametres.sqlite')
        cur = connection.cursor()
        if not cur:
            QMessageBox.critical(
                None,
                "App Name - Error!",
                "Database Error: %s" % con.lastError().databaseText(),
            )
        try:
            tablerow = 0
            # self.tableWidget.setRowCount(10)
            tablename = "Horaires"
            sqlstr = "SELECT COUNT(Jour_Date) FROM {} ".format(tablename)
            cur.execute(sqlstr)
            fixture_count = cur.fetchone()[0]
            # print(fixture_count)
            self.tableWidget.setRowCount(fixture_count)
            self.tableWidget.setColumnCount(10)
            sqlstr = "SELECT Jour_Date, Jour, Mois_AR, Mois_FR,heure_fajr, heure_shourouq, heure_dhuhr,heure_asr, " \
                     "heure_maghrib, heure_ishae FROM {} ".format(tablename)
            sqlstr = "SELECT * FROM {} ".format(tablename)
            cur.execute(sqlstr)
            result = cur.fetchall()
            # number_of_rows = len(result)
            # print("number_of_rows: ", number_of_rows)
            # fetch all the matching rows
            # loop through the rows
            tablerow = 0
            for row in result:
                # print(row)
                # print("\n")
                id = row[0]
                item = QtWidgets.QTableWidgetItem(row[0])
                item0 = QtWidgets.QTableWidgetItem(row[1])  # row 1
                item1 = QtWidgets.QTableWidgetItem(row[2])
                item2 = QtWidgets.QTableWidgetItem(row[3])
                item3 = QtWidgets.QTableWidgetItem(row[4])
                item4 = QtWidgets.QTableWidgetItem(row[5])
                item5 = QtWidgets.QTableWidgetItem(row[6])
                item6 = QtWidgets.QTableWidgetItem(row[7])
                item7 = QtWidgets.QTableWidgetItem(row[8])
                item8 = QtWidgets.QTableWidgetItem(row[9])
                item9 = QtWidgets.QTableWidgetItem(row[10])
                # item10= QtWidgets.QTableWidgetItem(row[11])
                self.tableWidget.setItem(tablerow, 0, item0)
                self.tableWidget.setItem(tablerow, 1, item1)
                self.tableWidget.setItem(tablerow, 2, item2)
                self.tableWidget.setItem(tablerow, 3, item3)
                self.tableWidget.setItem(tablerow, 4, item4)
                self.tableWidget.setItem(tablerow, 5, item5)
                self.tableWidget.setItem(tablerow, 6, item6)
                self.tableWidget.setItem(tablerow, 7, item7)
                self.tableWidget.setItem(tablerow, 8, item8)
                self.tableWidget.setItem(tablerow, 9, item9)
                # self.tableWidget.setItem(tablerow, 10, item10)
                tablerow = tablerow + 1
                # print("tablerow: ", tablerow)
                Jour_Date = row[1]
                Jour = row[2]
                Mois_AR = row[3]
                Mois_FR = row[4]
                heure_fajr = row[5]
                heure_shourouq = row[6]
                heure_dhuhr = row[7]
                heure_asr = row[8]
                heure_maghrib = row[9]
                heure_ishae = row[10]
                # print(id)
                # print(Jour_Date)
                # print(Jour)
                # print(Mois_AR)
                # print(Mois_FR)
                # print(heure_fajr)
                # print(heure_shourouq)
                # print(heure_dhuhr)
                # print(heure_asr)
                # print(heure_maghrib)
                # print(heure_ishae)
            cur.close()
            connection.close()
            self.btn_Valider.hide()
            self.btn_Annuler.hide()
            if not enModeTest():
                self.supprimer.hide()
                self.btn_supprimer.hide()
        except Exception as e:
            print("Exception ", e)
            traceback.print_tb(exc_traceback, limit=1, file=sys.stdout)

    QtCore.pyqtSlot()

    def closeEvent(self, evnt):
        print("closed evt")
        w = None
        self.signal.emit("CLOSED")

    QtCore.pyqtSlot()

    def akcja2(self):
        self.signal.emit("CLOSED")

    def openFile(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "Load Excel", '',
                                                  "Excel File(*.xlsx)")
        if fileName != "":
            self.tableWidget.setRowCount(0)
            workbook = openpyxl.load_workbook(fileName)
            sheet = workbook.active
            print("max_row", sheet.max_row)
            print("max_column", sheet.max_column)
            if sheet.max_row > 0:
                self.tableWidget.setRowCount(sheet.max_row - 1)
                self.tableWidget.setColumnCount(sheet.max_column)
                if sheet.values:
                    print("OK")
                else:
                    print("NG")
                print(sheet.values)

                bok = True
                if bok:
                    list_values = list(sheet.values)
                    self.tableWidget.setHorizontalHeaderLabels(list_values[0])

                    row_index = 0
                    for value_tuple in list_values[1:]:
                        col_index = 0
                        for value in value_tuple:
                            if col_index == 0:
                                value = value.strftime("%Y-%m-%d")
                            self.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                            col_index += 1
                        row_index += 1
                    # count item
                    rowCount = self.tableWidget.rowCount()
                    if rowCount > 0:
                        self.btn_Valider.show()
                        self.btn_Annuler.show()

    def accept(self):
        print("OK")

    def show_popup(self, count):
        msg = QMessageBox(self)
        msg.setWindowTitle("Importation Horaires")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.buttonClicked.connect(self.popup)
        pixmap = QPixmap('1431359280_ic_list_48px-24.png')
        # msg.setIconPixmap(pixmap)
        # icon = QIcon("png/1431359280_ic_list_48px-24.png")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("png/1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        msg.setWindowIcon(icon)
        if count == 0:
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Les Enregistrements Déja existe dans la base")
        else:
            msg.setIcon(QMessageBox.Information)
            msg.setText("Fichier bien enregistrer dans la base")

        x = msg.exec_()
        if x == QMessageBox.Ok:
            print("OK!")
        # msg.setDetailedText("Extra details.....")
        #     # msg.setText("This is some random text")
        #     # msg.setInformativeText("This is some extra informative text")
        #
        # x = msg.exec_()
        #

    def popup(self, i):
        print(i.text())

    def valider(self):
        global heure_fajr, heure_shourouq, heure_asr, heure_maghrib, heure_ishae, Mois_FR, Mois_AR, Jour, Jour_Date, heure_dhuhr
        print("Updating horaires...")
        bok = False
        exist = False
        count = 0
        no_error = True
        db_connection = sqlite3.connect('Parametres.sqlite')
        curs = db_connection.cursor()
        try:

            for row in range(self.tableWidget.rowCount()):
                bok = False
                exist = False
                no_error = True
                for col in range(self.tableWidget.columnCount()):
                    bok = False
                    exist = False
                    no_error = True
                    # it = self.tableWidget.item(row, 0)
                    # jar = self.tableWidget.item(i, 0).text()
                    widegetItem = self.tableWidget.item(row, col)
                    if widegetItem and widegetItem.text():
                        # 2023 - 01 - 25
                        ID_Date = self.tableWidget.item(row, 0).data(Qt.EditRole)
                        check_if_exist = "SELECT * FROM horaires WHERE Jour_Date='" + ID_Date + "'"
                        curs.execute(check_if_exist)
                        data_tmp = curs.fetchall()
                        #
                        exist = False
                        if len(data_tmp) == 0:
                            exist = False
                        else:
                            exist = True
                        print("exist" if exist else "not exist")
                        if not exist:
                            if not exist:
                                try:
                                    self.tableWidget.item(row, col).data(Qt.EditRole)
                                    no_error = True
                                except Exception as e:
                                    no_error = False
                                    print("Exception: ", e)
                                    traceback.print_tb(exc_traceback, limit=1, file=sys.stdout)
                                else:
                                    if no_error:
                                        if col == 0:
                                            Jour_Date = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Date Du Jour: ", Jour_Date)
                                        if col == 1:
                                            Jour = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Le Jour: ", Jour)
                                        if col == 2:
                                            Mois_AR = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Mois Arabe: ", Mois_AR)
                                        if col == 3:
                                            Mois_FR = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Mois Fr: ", Mois_FR)
                                        if col == 4:
                                            heure_fajr = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure fajr: ", heure_fajr)
                                        if col == 5:
                                            heure_shourouq = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure chourouq: ", heure_shourouq)
                                        if col == 6:
                                            heure_dhuhr = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure duhr: ", heure_dhuhr)
                                        if col == 7:
                                            heure_asr = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure Asr: ", heure_asr)
                                        if col == 8:
                                            heure_maghrib = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure Maghrib: ", heure_maghrib)
                                        if col == 9:
                                            heure_ishae = self.tableWidget.item(row, col).data(Qt.EditRole)
                                            print("Heure iSha: ", heure_ishae)
                                        if col == 9:
                                            bok = True

                                        else:
                                            bok = False

                                        if bok and not exist and no_error:
                                            count = count + 1
                                            sql = ''' INSERT INTO Horaires (Jour_Date, Jour, Mois_AR, Mois_FR,heure_fajr, heure_shourouq, heure_dhuhr,heure_asr, heure_maghrib, heure_ishae)
                                                         VALUES(?,?,?,?,?,?,?,?,?,?) '''

                                            curs.execute(sql, (
                                                Jour_Date, Jour, Mois_AR, Mois_FR, heure_fajr, heure_shourouq,
                                                heure_dhuhr, heure_asr,
                                                heure_maghrib, heure_ishae))
                                            print(curs.lastrowid)
                                            db_connection.commit()

                    else:
                        print('Cell is empty', row)
                        bok = False

                    # insertDataQuery = QSqlQuery()
                    # insertDataQuery.prepare(
                    #     """
                    #      INSERT INTO Horaires (
                    #         Jour_Date,
                    #         Jour,
                    #         Mois_AR,
                    #         Mois_FR ,
                    #         heure_fajr,
                    #         heure_shourouq,
                    #         heure_dhuhr ,
                    #         heure_asr TIME,
                    #         heure_maghrib ,
                    #         heure_ishae
                    #     )
                    #       VALUES (?, ?, ?,?, ?, ?,?, ?, ?,?)
                    #       """
                    # )
                    # Use .addBindValue() to insert data
                    # for name, job, email in data:
                    #     insertDataQuery.addBindValue(name)
                    #     insertDataQuery.addBindValue(job)
                    #     insertDataQuery.addBindValue(email)
                    #     insertDataQuery.exec_()

            # msg = QMessageBox()
            # msg.setWindowTitle("Message Box")
            # msg.setDefaultButton(QMessageBox.Ok)
            self.show_popup(count)
            # if count == 0:
            #
            #     # QMessageBox.critical(
            #     #     None,
            #     #     "Prayer Time - Information!",
            #     #     "Les Horaires déja Existe dans la base",
            #     # )
            #     # msg.setIcon(QMessageBox.informative)
            #     # msg.setDetailedText("Extra details.....")
            #     # msg.setText("This is some random text")
            #     # msg.setInformativeText("This is some extra informative text")
            # else:
            #     QMessageBox.information(
            #         None,
            #         "Prayer Time - Information!",
            #         "Fichier bien enregistrer dans la base"
            #     )
            #     msg.setIcon(QMessageBox.warning)
            #     msg.setDetailedText("Extra details.....")
            #     msg.setText("This is some random text")
            #     msg.setInformativeText("This is some extra informative text")
            # returnValue = msg.exec_()
            # if returnValue == QMessageBox.Ok:
            #     print('OK clicked')
            self.loaddata()
        except Exception as e:
            print("Exception: ", e)
            traceback.print_tb(exc_traceback, limit=1, file=sys.stdout)
            pass


# Note that you need to run `QSqlDatabase().commit()`` if you want the data to be committed in the database.
class Parametrages_Window(QDialog):
    signal = QtCore.pyqtSignal(str)

    def __init__(self):
        # super(Parametrages_Window, self).__init__()
        super().__init__()
        loadUi("ui/Parametrages.ui", self)
        self.w = self
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("png/1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        self.setWindowTitle("Parametrages")
        self.tableWidget.setColumnWidth(0, 100)
        self.tableWidget.setColumnWidth(1, 400)
        self.tableWidget.setColumnWidth(2, 412)
        self.loaddata()
        # create "global" list to save all selected items
        self.selected_items = []
        self.cell = []
        self.row, self.col = 0, 0
        self.Param_ID = 0
        self.Param_Nom = ""
        self.Param_Value = ""
        self.bok = False
        self.buttonBox.clicked.connect(self.clickme)
        # ±self.tableWidget.setGeometry(QtCore.QRect(0, 0, 552, 331))
        self.tableWidget.keyPressEvent = self.KeyPressed
        self.tableWidget.itemChanged.connect(self.save_changes)
        self.tableWidget.itemSelectionChanged.connect(self.print_row)
        self.tableWidget.cellClicked.connect(self.cellClick)
        self.tableWidget.cellChanged['int', 'int'].connect(self.cellChanged)
        self.tableWidget.cellChanged['int', 'int'].connect(lambda: self.accept)
        # self.tableWidget.cellPressed['int','int'].connect(self.KeyPressed) # type:
        # self.tableWidget.itemClicked.connect(self.onClicked)
        # self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.OK.clicked.connect(self.accept) # type: ignore

    def open_second(self):
        print("open second window")

    #         self.widget.setWindowTitle("Horaires")
    #         icon = QtGui.QIcon()
    #         icon.addPixmap(QtGui.QPixmap("1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
    #         self.widget.setWindowIcon(icon)
    #         #self.widget.resize(550, 412)
    #         self.widget.resize(550, 372)
    #         self.widget.setMaximumSize(QtCore.QSize(550, 372))
    #         self.widget.show()

    def cellClick(self, row, col):
        self.row = row
        self.col = col

    def print_row(self):
        items = self.tableWidget.selectedItems()
        for currentQTableWidgetItem in items:
            print(str(items[0].text()))

    def cellChanged(self, row, col):
        # The first in green (0,255,0)
        # Other in red (255,0,0)
        self.row = row
        self.col = col
        items = self.tableWidget.selectedItems()
        for currentQTableWidgetItem in items:
            pass
            # print("selection changed: ",str(items[0].text()))
        self.buttonBox.show()

    # action method
    def clickme(self):
        # hiding the button
        self.buttonBox.hide()
        # printing pressed
        print("pressed")

    @QtCore.pyqtSlot(QtWidgets.QTableWidgetItem)
    def onClicked(self, it):
        state = not it.data(SelectedRole)
        it.setData(SelectedRole, state)
        it.setBackground(
            QtGui.QColor(100, 100, 100) if state else QtGui.QColor(0, 255, 127)
        )

    def loaddata(self):
        connection = sqlite3.connect('Parametres.sqlite')
        cur = connection.cursor()
        if not cur:
            QMessageBox.critical(
                None,
                "App Name - Error!",
                "Database Error: %s" % con.lastError().databaseText(),
            )
        tablerow = 0
        self.tableWidget.setRowCount(tablerow)
        self.tableWidget.setRowCount(11)
        tablename = "Parametres"
        sqlstr = "SELECT * FROM {} ".format(tablename)
        results = cur.execute(sqlstr)
        for row in results:
            item1 = QtWidgets.QTableWidgetItem(row[0])
            item2 = QtWidgets.QTableWidgetItem(row[1])
            item3 = QtWidgets.QTableWidgetItem(row[2])
            # item.setFlags(QtCore.Qt.ItemIsSelectable|QtCore.Qt.ItemIsEditable|QtCore.Qt.ItemIsDragEnabled|QtCore.Qt.ItemIsUserCheckable|QtCore.Qt.ItemIsEnabled)
            if not tablerow == 11 or tablerow == 12:
                self.tableWidget.setItem(tablerow, 0, item1)
                self.tableWidget.setItem(tablerow, 0, item2)
                self.tableWidget.setItem(tablerow, 1, item3)
            # item2.setFlags(Qt.ItemIsSelectable |  Qt.ItemIsEnabled)
            item2.setFlags(item2.flags() & ~QtCore.Qt.ItemIsEditable)
            if tablerow == 0 or tablerow == 1 or tablerow == 2 or tablerow == 7 or tablerow == 8:
                item3.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            #                 it.setFlags(it.flags() & ~Qt.ItemIsSelectable)
            tablerow += 1
        #             if tablerow == 10 or tablerow == 11:
        #                 self.tableWidget.hideRow(tablerow);
        cur.close()
        connection.close()

    @pyqtSlot()
    # @QtCore.pyqtSlot(QtWidgets.QTableWidgetItem)
    def save_changes(self):
        self.buttonBox.show()
        for currentQTableWidgetItem in self.tableWidget.selectedItems():
            self.Param_ID = self.tableWidget.item(currentQTableWidgetItem.row(), 0).text()
            self.Param_Nom = self.tableWidget.item(currentQTableWidgetItem.row(), 0).text()
            self.Param_Value = self.tableWidget.item(currentQTableWidgetItem.row(), 1).text()

    def KeyPressed(self, event):
        if event.key() == QtCore.Qt.Key_Return:
            print('Enter Key Pressed')
            self.bok = True
            for currentQTableWidgetItem in self.tableWidget.selectedItems():
                self.Param_ID = self.tableWidget.item(currentQTableWidgetItem.row(), 0).text()
                self.Param_Nom = self.tableWidget.item(currentQTableWidgetItem.row(), 0).text()
                self.Param_Value = self.tableWidget.item(currentQTableWidgetItem.row(), 1).text()

            if self.Param_Nom == "remining_Time":
                remining_Time = self.Param_Value
                self.bok = isTimeFormat(remining_Time)
            if not self.bok:
                QMessageBox.critical(None, "App Name - Error!", "Champ Invalide: %s" % remining_Time, )
                for currentQTableWidgetItem in self.tableWidget.selectedItems():
                    self.tableWidget.item(currentQTableWidgetItem.row(), 1).setBackground(QtGui.QColor(255, 0, 0))
            else:
                for currentQTableWidgetItem in self.tableWidget.selectedItems():
                    self.tableWidget.item(currentQTableWidgetItem.row(), 1).setBackground(QtGui.QColor(0, 255, 0))

            if not self.bok:
                QMessageBox.critical(None, "App Name - Error!", "Champ Invalide: %s" % self.Param_Nom, )
            if self.bok:
                print("Param_ID", self.Param_ID)
                print("Param_Nom", self.Param_Nom)
                print("Param_Value", self.Param_Value)
                file = pathlib.Path("config/config.ini")
                if not file.exists():
                    print("Configuration File not exist")
                else:
                    parser = configparser.ConfigParser()
                    parser.read('config/config.ini')
                    s_Param_Nom = self.Param_Nom.lower()
                    s_Param_Value = self.Param_Value
                    parser.set('DEFAULT', s_Param_Nom, s_Param_Value)
                    # Writing our configuration file to 'example.ini'
                    with open('config/config.ini', 'w') as configfile:
                        parser.write(configfile)

                connection = sqlite3.connect('Parametres.sqlite')
                cur = connection.cursor()
                sqlstr = "UPDATE Parametres SET Param_Valeur='%s' WHERE Param_Nom='%s'" % (
                    self.Param_Value, self.Param_Nom)
                cur.execute(sqlstr)
                connection.commit()
                cur.close()
                connection.close()
                self.loaddata()
                self.buttonBox.hide()
                QMessageBox.about(self, "info", "Table Parametres Updated")
                if self.Param_Nom == "ResolutionW" or self.Param_Nom == "ResolutionH" or self.Param_Nom == "city" or self.Param_Nom == "remining_Time" or self.Param_Nom == "mode":
                    os.execv(sys.executable, [sys.executable] + sys.argv)

    def get_colour(self):
        if self.type == CorrectionBox.types.TO_CORRECT:
            return QtGui.QColor(255, 0, 0)
        elif self.type == CorrectionBox.types.TO_REVIEW:
            return QtGui.QColor(255, 255, 0)
        elif self.type == CorrectionBox.types.RESOLVED:
            return QtGui.QColor(0, 255, 0)
        elif self.type == CorrectionBox.types.QUESTION:
            return QtGui.QColor(0, 0, 255)

    def reject(self):
        ## The first in green (0,255,0)
        # Other in red (255,0,0)

        for currentQTableWidgetItem in self.tableWidget.selectedItems():
            self.tableWidget.item(currentQTableWidgetItem.row(), 1).setBackground(QtGui.QColor(255, 255, 0))
        print("reject")
        self.buttonBox.hide()

    def accept(self):

        self.bok = False
        remining_Time = self.Param_Value
        if self.Param_Nom == "remining_Time":
            remining_Time = self.Param_Value
            self.bok = isTimeFormat(remining_Time)
        if not self.bok:
            QMessageBox.critical(None, "App Name - Error!", "Champ Invalide: %s" % remining_Time, )
            for currentQTableWidgetItem in self.tableWidget.selectedItems():
                self.tableWidget.item(currentQTableWidgetItem.row(), 1).setBackground(QtGui.QColor(255, 0, 0))
        else:
            for currentQTableWidgetItem in self.tableWidget.selectedItems():
                self.tableWidget.item(currentQTableWidgetItem.row(), 1).setBackground(QtGui.QColor(0, 255, 0))

        if not self.bok:
            QMessageBox.critical(None, "App Name - Error!", "Champ Invalide: %s" % self.Param_Nom, )
        if self.bok:
            print("Param_ID", self.Param_ID)
            print("Param_Nom", self.Param_Nom)
            print("Param_Value", self.Param_Value)
            file = pathlib.Path("config/config.ini")
            if not file.exists():
                print("Configuration File not exist")
            else:
                parser = configparser.ConfigParser()
                parser.read('config/config.ini')
                s_Param_Nom = self.Param_Nom.lower()
                s_Param_Value = self.Param_Value
                parser.set('DEFAULT', s_Param_Nom, s_Param_Value)
                # Writing our configuration file to 'example.ini'
                with open('config/config.ini', 'w') as configfile:
                    parser.write(configfile)
            #                 updater = ConfigUpdater()
            #                 updater.read("config.ini")
            #                 updater["DEFAULT"][self.Param_Nom.lower()]= self.Param_Value

            connection = sqlite3.connect('Parametres.sqlite')
            cur = connection.cursor()
            sqlstr = "UPDATE Parametres SET Param_Valeur='%s' WHERE Param_Nom='%s'" % (self.Param_Value, self.Param_Nom)
            cur.execute(sqlstr)
            connection.commit()
            cur.close()
            connection.close()
            self.loaddata()
            self.buttonBox.hide()
            QMessageBox.about(self, "info", "Table Parametres Updated")
            if self.Param_Nom == "ResolutionW" or self.Param_Nom == "ResolutionH" or self.Param_Nom == "city" or self.Param_Nom == "remining_Time" or self.Param_Nom == "mode":
                os.execv(sys.executable, [sys.executable] + sys.argv)

    def akcja2(self):
        self.signal.emit("CLOSED")

    def closeEvent(self, evnt):
        # print(QSqlDatabase.connectionNames())
        # Remove the default connection
        self.signal.emit("CLOSED")
        print("Close Parametre")
        QSqlDatabase.removeDatabase(QSqlDatabase.database().connectionName())

    def setenable(self):
        self.OK.setEnabled(True)


class MainWindow(QMainWindow):
    my_signal = QtCore.pyqtSignal(int)

    def __init__(self):
        # super(MainWindow, self).__init__() # Call the inherited classes __init__ method
        super().__init__()
        uic.loadUi('ui/basic.ui', self)  # Load the .ui file
        self.w = None  # No external window yet.
        self.mode = mode
        self.debug = debug
        self.windows = []
        #         app_icon.addFile('gui/icons/16x16.png', QtCore.QSize(16,16))
        #         app_icon.addFile('gui/icons/24x24.png', QtCore.QSize(24,24))
        #         app_icon.addFile('gui/icons/32x32.png', QtCore.QSize(32,32))
        #         app_icon.addFile('gui/icons/48x48.png', QtCore.QSize(48,48))
        #         app_icon.addFile('gui/icons/256x256.png', QtCore.QSize(256,256))
        self.setWindowIcon(QtGui.QIcon('icon/icon.png'))
        # self.showFullScreen()
        self.setWindowTitle('Prayer Times')
        # url(DSC_0001.jpg)
        stylesheet = '''
    #MainWindow {
        background-image: url(''' + background_image_url + ''');
        background-repeat: no-repeat;
        background-position: center;
    }
'''
        self.setStyleSheet(stylesheet)
        self.label_fajr.setPixmap(QtGui.QPixmap(fajr_image))
        self.label_dhuhr.setPixmap(QtGui.QPixmap(dhuhr_image))
        self.label_asr.setPixmap(QtGui.QPixmap(asr_image))
        self.label_maghrib.setPixmap(QtGui.QPixmap(maghrib_image))
        self.label_isha.setPixmap(QtGui.QPixmap(isha_image))

        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("png/1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.menuParametrages.setTitle("Menu")
        self.menuParametrages.setIcon(icon)
        # self.menuParametrages.setIcon(UI.PixmapCache.getIcon("1431359280_ic_list_48px-24.png"))
        self.statusBar().showMessage('Prayer Time 2023 - MNE')
        # 2700x1700
        # 2048x1365
        # 1920x1200
        # 1600x1200
        # 1366 x768
        self.resize(int(ResolutionW), int(ResolutionH))
        self.window_width, self.window_height = int(ResolutionW), int(ResolutionH)
        self.setMinimumSize(self.window_width, self.window_height)
        self.setWindowIcon(QtGui.QIcon('png/logo.png'))

        self.setMinimumSize(QtCore.QSize(int(ResolutionW), int(ResolutionH)))
        self.setMaximumSize(QtCore.QSize(int(ResolutionW), int(ResolutionH)))
        # load ini config
        #         font_db = QFontDatabase()
        #         font_id = font_db.addApplicationFont('Othmani.ttf')
        # self.LIB_TIPS.setFont(QFont('Othmani'))
        # self.LIB_TIPS.setStyleSheet("QLabel{font-family:'Othmani'}")
        # self.LIB_TIPS.setText(note_Journee)
        self.city = city
        self.timer = QTimer()
        self.timer.start(1000)
        self.timer.timeout.connect(self.thread)

        # create media player object
        # self.mediaPlayer = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        self.player = QMediaPlayer()
        self.player.stateChanged.connect(self.mediastate_changed)
        self.player.stateChanged.connect(self.thread)
        currentVolume = self.player.volume()  #
        self.started = 0
        self.timerupdate = QTimer()
        self.timerupdate.start(10000)
        if self.mode == "Online":
            self.showTimeprayer()
            self.timerupdate.timeout.connect(self.threadshowTimeprayer)

        # set title
        self.showTiTle()
        self.actionParametrages.triggered.connect(self.open_settings)
        self.actionHoraires.triggered.connect(self.open_Horaires)
        #         Screen1 = Parametrages_Window()
        #
        #         self.widget = QtWidgets.QStackedWidget()
        #         self.widget.addWidget(Screen1)

        #         self.widget.setFixedHeight(550)
        #         self.widget.setFixedWidth(580)
        # Set geometry
        self.show()  # Show the GUI

    def slot1(self, X):
        print("X is clicked", X)
        self.w = None

    #     def openFile(self):
    #         fileName, _ = QFileDialog.getOpenFileName(self, "Load Excel",QDir.homePath())
    #         print(fileName)
    #         return

    def thread(self):
        self.t1 = Thread(target=self.showTime)
        self.t1.start()

    def threadshowTimeprayer(self):
        t2 = Thread(target=self.showTimeprayer)
        t2.start()

    def showTime(self):
        time = QDateTime.currentDateTime()
        timeDisplay = time.toString('yyyy-MM-dd hh:mm:ss')
        if not mode == "Online":
            self.LIB_DT.setText(timeDisplay)
        url_api = MONURL1 + self.city + ".html"
        # print(url_api)
        # print("Mode: ", mode)
        if mode == "Online":
            try:
                if not self.player.state() == QMediaPlayer.PlayingState:
                    if self.debug: print("is stopped", remining_Time)
                    r = requests.get(url_api, timeout=int(time_out))
                    if r.status_code == 200:
                        soup = BeautifulSoup(r.content, 'html.parser')
                        for data in soup(['style', 'script']):
                            data.decompose()
                            # Remove tags
                        reminingTime = soup.find('div', class_="reminingTime")
                        # print("reminingTime",reminingTime.getText())
                        j = 0
                        for i in reminingTime:
                            # print(i)
                            j = j + 1
                            if j == 1:
                                kel_priere = i.extract().getText()
                                self.LIB_DT_2.setText(kel_priere)
                                # print()
                        countdown = soup.find('div', id="countdown").getText()
                        self.LIB_remining_Time.setText(countdown)
                        print("countdown: ",countdown)
                        # countdown
                        # print(kel_priere)
                        heure_sys = datetime.datetime.now().strftime("%H:%M")
                        # print(heure_sys)
                        if kel_priere == "Fajr":
                            lheure_fajr = self.LIB_heure_fajr.text()
                            heure_azan = lheure_fajr
                        if kel_priere == "Dhuhr":
                            lheure_dhuhr = self.LIB_heure_dhuhr.text()
                            heure_azan = lheure_dhuhr
                        if kel_priere == "Asr":
                            # Vérifier l'heure asr sinon remaining
                            lheure_asr = self.LIB_heure_asr.text()
                            heure_azan = lheure_asr
                            # print(heure_asr)
                        if kel_priere == "Maghrib":
                            lheure_maghrib = self.LIB_heure_maghrib.text()
                            heure_azan = lheure_maghrib
                        if kel_priere == "Isha":
                            lheure_isha = self.LIB_heure_ishae.text()
                            heure_azan = lheure_isha
                        if heure_sys == heure_azan or countdown == "00:00:00":
                            # time to pray
                            self.LIB_DT_4.setText("حان وقت الاذان")
                            self.LIB_remining_Time.setText(heure_azan)
                            self.timer.stop()
                            self.playAudioFile()
                    else:
                        print("Mode Offline")
                        # definir la période en se basant sur l'heure
                        heure_actuel = datetime.datetime.now().strftime("%H:%M:%S")
                        bheure_actuel = parser.parse(heure_actuel)
                        lheure_actuel = bheure_actuel.strftime("%H:%M:%S")
                        # print("heure actuel: ", lheure_actuel)
                        if self.debug: print("heure_actuel>>", lheure_actuel)
                        # kel journée
                        kel_Date = datetime.datetime.now().strftime("%Y-%m-%d")
                        # Récupérer les horaires tel date
                        new_gb_heure_fajr = get_Horaire(kel_Date, "heure_fajr")
                        gb_heure_fajr = new_gb_heure_fajr
                        c = 0
                        if new_gb_heure_fajr is None:
                            print("Non Trouvé")
                            bgo = False
                            c = 0
                        else:
                            bgo = True
                            self.LIB_heure_fajr.setText(gb_heure_fajr[:5])
                            c = 1

                        gb_heure_shourouq = get_Horaire(kel_Date, "heure_shourouq")
                        if gb_heure_shourouq is None:
                            if self.debug: print("Non Trouvé")
                            bgo = False
                            c = c - 1
                        else:
                            bgo = True
                            c = c + 1

                        gb_heure_dhuhr = get_Horaire(kel_Date, "heure_dhuhr")
                        if gb_heure_dhuhr is None:
                            if self.debug: print("Non Trouvé")
                            bgo = False
                            c = c - 1
                        else:
                            bgo = True
                            self.LIB_heure_dhuhr.setText(gb_heure_dhuhr[:5])
                            c = c + 1

                        gb_heure_asr = get_Horaire(kel_Date, "heure_asr")
                        if gb_heure_asr is None:
                            if self.debug: print("Non Trouvé")
                            bgo = False
                            c = c - 1
                        else:
                            bgo = True
                            self.LIB_heure_asr.setText(gb_heure_asr[:5])
                            c = c + 1

                        gb_heure_maghrib = get_Horaire(kel_Date, "heure_maghrib")
                        if gb_heure_maghrib is None:
                            if self.debug: print("Non Trouvé")
                            bgo = False
                            c = c - 1
                        else:
                            bgo = True
                            c = c + 1
                            self.LIB_heure_maghrib.setText(gb_heure_maghrib[:5])

                        gb_heure_ishae = get_Horaire(kel_Date, "heure_ishae")
                        if gb_heure_ishae is None:
                            if self.debug: print("Non Trouvé")
                            bgo = False
                            c = c - 1
                        else:
                            bgo = True
                            c = c + 1
                            self.LIB_heure_ishae.setText(gb_heure_ishae[:5])

                        if bgo:
                            if debug: print("c: ", c)
                            if c != 6:
                                bgo = False

                        if bgo:
                            # calcul difference
                            if new_gb_heure_fajr < heure_actuel <= gb_heure_dhuhr:
                                self.LIB_DT_2.setText("Dhuhr")
                                if gb_heure_dhuhr == heure_actuel:
                                    if self.debug: print("c'est l'heure dhuhr")
                                    self.LIB_DT_4.setText("حان وقت الاذان")
                                    self.playAudioFile()
                                else:
                                    bheure_dhuhr = parser.parse(gb_heure_dhuhr)
                                    lheure_dhuhr = bheure_dhuhr.strftime("%H:%M:%S")
                                    print("Heure dhuhr: ", lheure_dhuhr)
                                    time_diff = (parser.parse(lheure_dhuhr) - parser.parse(lheure_actuel))
                                    tsecs = time_diff.total_seconds()
                                    tmins = tsecs / 60
                                    thrs = tsecs / (60 * 60)
                                    res = datetime.timedelta(seconds=tsecs)
                                    print("Temps_Restant>> ", res)
                                    Temps_Restant = str(res).zfill(8)
                                    self.LIB_remining_Time.setText(Temps_Restant)

                            # tester si l'heure actuel < heure asr
                            if gb_heure_dhuhr < heure_actuel <= gb_heure_asr:
                                print("l'heure asr")
                                self.LIB_DT_2.setText("Asr")
                                if gb_heure_asr == heure_actuel:
                                    print("c'est l'heure asr")
                                    self.LIB_DT_4.setText("حان وقت الاذان")
                                    self.playAudioFile()
                                else:
                                    # calcul difference to asr
                                    bheure_asr = parser.parse(gb_heure_asr)
                                    lheure_asr = bheure_asr.strftime("%H:%M:%S")
                                    print("Heure asr: ", lheure_asr)
                                    time_diff = (parser.parse(lheure_asr) - parser.parse(lheure_actuel))
                                    tsecs = time_diff.total_seconds()
                                    tmins = tsecs / 60
                                    thrs = tsecs / (60 * 60)
                                    res = datetime.timedelta(seconds=tsecs)
                                    print("Temps_Restant>> ", res)
                                    Temps_Restant = str(res).zfill(8)
                                    self.LIB_remining_Time.setText(Temps_Restant)

                            if gb_heure_asr < heure_actuel <= gb_heure_maghrib:
                                # heure el maghrib
                                self.LIB_DT_2.setText("Maghrib")
                                if gb_heure_maghrib == heure_actuel:
                                    if self.debug: print("c'est l'heure maghrib")
                                    self.LIB_DT_4.setText("حان وقت الاذان")
                                    self.playAudioFile()
                                else:
                                    bheure_maghrib = parser.parse(gb_heure_maghrib)
                                    lheure_maghrib = bheure_maghrib.strftime("%H:%M:%S")
                                    print("Heure maghrib: ", lheure_maghrib)
                                    # time_diff = lheure_actuel - lheure_maghrib
                                    time_diff = (parser.parse(lheure_maghrib) - parser.parse(lheure_actuel))
                                    tsecs = time_diff.total_seconds()
                                    tmins = tsecs / 60
                                    thrs = tsecs / (60 * 60)
                                    res = datetime.timedelta(seconds=tsecs)
                                    print("Temps_Restant>> ", res)
                                    Temps_Restant = str(res).zfill(8)
                                    self.LIB_remining_Time.setText(Temps_Restant)

                            if gb_heure_maghrib < heure_actuel <= gb_heure_ishae:

                                self.LIB_DT_2.setText("Isha")

                                if gb_heure_ishae == heure_actuel:
                                    if self.debug: print("c'est l'heure ishae")
                                    self.LIB_DT_4.setText("حان وقت الاذان")
                                    self.playAudioFile()
                                else:
                                    bheure_ishae = parser.parse(gb_heure_ishae)
                                    lheure_ishae = bheure_ishae.strftime("%H:%M:%S")
                                    print("Heure ishae: ", lheure_ishae)
                                    time_diff = (parser.parse(lheure_ishae) - parser.parse(lheure_actuel))
                                    tsecs = time_diff.total_seconds()
                                    tmins = tsecs / 60
                                    thrs = tsecs / (60 * 60)
                                    res = datetime.timedelta(seconds=tsecs)
                                    print("Temps_Restant>> ", res)
                                    Temps_Restant = str(res).zfill(8)
                                    self.LIB_remining_Time.setText(Temps_Restant)

                            if gb_heure_ishae < heure_actuel:

                                # Récupération Fajr next day

                                curr_date = datetime.datetime.now().strftime("%Y-%m-%d")
                                today = parser.parse(curr_date + " " + heure_actuel)
                                curr_date_temp = datetime.datetime.strptime(curr_date, "%Y-%m-%d")
                                new_date = curr_date_temp + datetime.timedelta(days=1)
                                new_date = new_date.strftime("%Y-%m-%d")
                                if self.debug: print("new date: ", new_date)
                                new_gb_heure_fajr = get_Horaire(new_date, "heure_fajr")
                                self.LIB_heure_fajr.setText(new_gb_heure_fajr[:5])
                                new_date_heure_fajr = new_date + " " + new_gb_heure_fajr
                                bday = parser.parse(new_date_heure_fajr)
                                if self.debug: print("New time: ", new_date_heure_fajr)
                                self.LIB_DT_2.setText("Fajr")
                                if new_gb_heure_fajr == heure_actuel:
                                    if self.debug: print("c'est l'heure el fajr")
                                    self.LIB_DT_4.setText("حان وقت الاذان")
                                    self.playAudioFile()
                                else:
                                    # calcul difference en heure entre les deux date
                                    time_diff = bday - today
                                    tsecs = time_diff.total_seconds()
                                    tmins = tsecs / 60
                                    thrs = tsecs / (60 * 60)
                                    res = datetime.timedelta(seconds=tsecs)
                                    print("Temps_Restant>> ", res)
                                    Temps_Restant = str(res).zfill(8)
                                    self.LIB_remining_Time.setText(Temps_Restant)
            except Exception as e:
                print("Exception: ", e)
        else:
            print("Mode Offline")
            # definir la période en se basant sur l'heure
            heure_actuel = datetime.datetime.now().strftime("%H:%M:%S")
            bheure_actuel = parser.parse(heure_actuel)
            lheure_actuel = bheure_actuel.strftime("%H:%M:%S")
            # print("heure actuel: ", lheure_actuel)
            if self.debug: print("heure_actuel>>", lheure_actuel)
            # kel journée
            kel_Date = datetime.datetime.now().strftime("%Y-%m-%d")
            # Récupérer les horaires tel date
            new_gb_heure_fajr = get_Horaire(kel_Date, "heure_fajr")
            gb_heure_fajr = new_gb_heure_fajr
            c = 0
            if new_gb_heure_fajr is None:
                print("Non Trouvé")
                bgo = False
                c = 0
            else:
                bgo = True
                self.LIB_heure_fajr.setText(gb_heure_fajr[:5])
                c = 1

            gb_heure_shourouq = get_Horaire(kel_Date, "heure_shourouq")
            if gb_heure_shourouq is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1

            gb_heure_dhuhr = get_Horaire(kel_Date, "heure_dhuhr")
            if gb_heure_dhuhr is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                self.LIB_heure_dhuhr.setText(gb_heure_dhuhr[:5])
                c = c + 1

            gb_heure_asr = get_Horaire(kel_Date, "heure_asr")
            if gb_heure_asr is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                self.LIB_heure_asr.setText(gb_heure_asr[:5])
                c = c + 1

            gb_heure_maghrib = get_Horaire(kel_Date, "heure_maghrib")
            if gb_heure_maghrib is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1
                self.LIB_heure_maghrib.setText(gb_heure_maghrib[:5])

            gb_heure_ishae = get_Horaire(kel_Date, "heure_ishae")
            if gb_heure_ishae is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1
                self.LIB_heure_ishae.setText(gb_heure_ishae[:5])

            if bgo:
                if debug: print("c: ", c)
                if c != 6:
                    bgo = False

            if bgo:
                # calcul difference
                if new_gb_heure_fajr < heure_actuel <= gb_heure_dhuhr:
                    self.LIB_DT_2.setText("Dhuhr")
                    if gb_heure_dhuhr == heure_actuel:
                        if self.debug: print("c'est l'heure dhuhr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_dhuhr = parser.parse(gb_heure_dhuhr)
                        lheure_dhuhr = bheure_dhuhr.strftime("%H:%M:%S")
                        print("Heure dhuhr: ", lheure_dhuhr)
                        time_diff = (parser.parse(lheure_dhuhr) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                # tester si l'heure actuel < heure asr
                if gb_heure_dhuhr < heure_actuel <= gb_heure_asr:
                    print("l'heure asr")
                    self.LIB_DT_2.setText("Asr")
                    if gb_heure_asr == heure_actuel:
                        print("c'est l'heure asr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        # calcul difference to asr
                        bheure_asr = parser.parse(gb_heure_asr)
                        lheure_asr   =bheure_asr.strftime("%H:%M:%S")
                        print("Heure asr: ",lheure_asr)
                        time_diff = (parser.parse(lheure_asr) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_asr < heure_actuel <= gb_heure_maghrib:
                    # heure el maghrib
                    self.LIB_DT_2.setText("Maghrib")
                    if gb_heure_maghrib == heure_actuel:
                        if self.debug: print("c'est l'heure maghrib")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_maghrib = parser.parse(gb_heure_maghrib)
                        lheure_maghrib = bheure_maghrib.strftime("%H:%M:%S")
                        print("Heure maghrib: ", lheure_maghrib)
                        # time_diff = lheure_actuel - lheure_maghrib
                        time_diff = (parser.parse(lheure_maghrib) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_maghrib < heure_actuel <= gb_heure_ishae:

                    self.LIB_DT_2.setText("Isha")

                    if gb_heure_ishae == heure_actuel:
                        if self.debug: print("c'est l'heure ishae")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_ishae = parser.parse(gb_heure_ishae)
                        lheure_ishae = bheure_ishae.strftime("%H:%M:%S")
                        print("Heure ishae: ", lheure_ishae)
                        time_diff = (parser.parse(lheure_ishae) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_ishae < heure_actuel:

                    # Récupération Fajr next day

                    curr_date = datetime.datetime.now().strftime("%Y-%m-%d")
                    today = parser.parse(curr_date + " " + heure_actuel)
                    curr_date_temp = datetime.datetime.strptime(curr_date, "%Y-%m-%d")
                    new_date = curr_date_temp + datetime.timedelta(days=1)
                    new_date = new_date.strftime("%Y-%m-%d")
                    if self.debug: print("new date: ", new_date)
                    new_gb_heure_fajr = get_Horaire(new_date, "heure_fajr")
                    self.LIB_heure_fajr.setText(new_gb_heure_fajr[:5])
                    new_date_heure_fajr = new_date + " " + new_gb_heure_fajr
                    bday = parser.parse(new_date_heure_fajr)
                    if self.debug: print("New time: ", new_date_heure_fajr)
                    self.LIB_DT_2.setText("Fajr")
                    if new_gb_heure_fajr == heure_actuel:
                        if self.debug: print("c'est l'heure el fajr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        # calcul difference en heure entre les deux date
                        time_diff = bday - today
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)
                        # if self.debug:
                        if True:
                            print(f"time_diff is {tsecs} seconds .")
                            print(f"time_diff is {tmins} minutes.")
                            print(f"time_diff  is {thrs} hours .")


    def mediastate_changed(self, state):
        if self.player.state() == QMediaPlayer.PlayingState:
            # self.playBtn.setIcon(
            #     self.style().standardIcon(QStyle.SP_MediaPause)
            #
            # )
            print("is playingg")
        else:
            print("is stopped")
            self.showTime()
            self.timer.start()
            # self.playBtn.setIcon(
            #     self.style().standardIcon(QStyle.SP_MediaPlay)
            #
            # )

    def handle_errors(self):
        pass
        # self.label.setText("Error: " + self.mediaPlayer.errorString())

    def playAudioFile(self):
        currentVolume = self.player.volume()  #
        if currentVolume != 100:
            self.player.setVolume(20)
        if not self.started:
            self.started = 1
            # CHECK IF FILE EXIST
            if check_if_existe(azan_mp3):
                full_file_path = os.path.join(os.getcwd(), azan_mp3)
                url = QUrl.fromLocalFile(full_file_path)
                content = QMediaContent(url)
                self.player.setMedia(content)
                self.player.play()

    def open_settings(self):
        if self.w is None:
            self.w = Parametrages_Window()
            self.w.signal.connect(self.slot1)
            self.windows.append(self.w)
        else:
            self.w = Parametrages_Window()
        self.w.show()

    #         self.widget.setWindowTitle("Parametres")
    #         icon = QtGui.QIcon()
    #         icon.addPixmap(QtGui.QPixmap("1431359280_ic_list_48px-24.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
    #         self.widget.setWindowIcon(icon)
    #         self.widget.resize(550, 372)
    #         self.widget.setMaximumSize(QtCore.QSize(550, 372))
    #         self.widget.setCurrentIndex(self.widget.currentIndex()+1)
    #         self.widget.show()

    def open_Horaires(self):
        if self.w is None:
            self.w = Horaires_Window()
            self.windows.append(self.w)
            self.w.signal.connect(self.slot1)
        else:
            self.w = Horaires_Window()

        self.w.show()

    def showTiTle(self):
        # traitement online
        # /mohammedia/
        MONURL2_old = "http://www.prayertimes.org/morocco"
        # url = MONURL2_old + "/" + self.city + "/"
        print(self.city)
        URL = "https://www.prayertimes.org/morocco/"+str(self.city)+"/"
        print("url: ",URL)
        if self.mode == "Online":
            try:
                # defining a params dict for the parameters to be sent to the API
                # PARAMS = {'city':  "morocco",'country':self.city}
                # URL = "https://www.prayertimes.org"
                # # sending get request and saving the response as response object
                # r = requests.get(url=URL, params=PARAMS)

                # extracting data in json format
                # data = r.json()
                r = requests.get(URL, timeout=5)
                if r.status_code != 200:
                    print("Failed to get Title from the server")
                else:
                    soup = BeautifulSoup(r.content, 'html.parser')
                    # Print the extracted data
                    for data in soup(['style', 'script']):
                        # Remove tags
                        data.decompose()
                    # replace with `soup.findAll` if you are using BeautifulSoup3
                    for div in soup.find_all('ins', class_="adsbygoogle"):
                        div.decompose()
                    title = soup.select('h1.prayer-times')[0].text.strip()
                    title = title.replace("وموعد الأذان", '')
                    self.LIB_INFO.setText(title)


            except Exception as e:
                print("Exception: ", e)
        else:
            # set title manuallty transtalte french to arabic
            if self.city =="mohammedia":
                city_AR =  "المحمدية"
            if self.city == "casablanca":
                city_AR = "الدار البيضاء"
            title = '''  مواقيت الصلاة في  {} المغرب 
            اليوم. الأوقات الرسمية للصلاة.'''
            self.LIB_INFO.setText(title.format(city_AR))

    def showTimeprayer(self):
        New_MONURL2 = "https://priere.ma/horaire-priere"
        url = New_MONURL2 + "-" + self.city
        if self.debug: print(url)
        mode = self.mode
        if self.mode == "Online":
            try:
                r = requests.get(url, timeout=int(time_out))
                if r.status_code != 200:
                    print("Failed to get horaires data from the server")
                    mode = "Offline"
                else:
                    mode="Online"
                    soup = BeautifulSoup(r.content, 'html.parser')
                    # Print the extracted data
                    for data in soup(['style', 'script']):
                        # Remove tags
                        data.decompose()

                    current_status = soup.find('span', class_="current-status").getText()
                    self.LIB_DT.setText(current_status)
                    # date_heure = soup.find(id="date_heure")
                    # print("date_heure: ",date_heure)
                    i = 0
                    for x in soup.find_all('span', class_="opening-hours-time"):
                        # fetching text from tag and remove whitespaces
                        i = i + 1
                        if self.debug: print(x.getText())
                        if i == 1:
                            gb_heure_fajr = x.getText()
                            self.LIB_heure_fajr.setText(gb_heure_fajr)
                        if i == 2:
                            pass
                            # self.LIB_heure_dhuhr.setText(x.getText())
                        if i == 3:
                            gb_heure_dhuhr = x.getText()
                            self.LIB_heure_dhuhr.setText(gb_heure_dhuhr)
                        if i == 4:
                            gb_heure_asr = x.getText()
                            self.LIB_heure_asr.setText(gb_heure_asr)
                        if i == 5:
                            gb_heure_maghrib = x.getText()
                            self.LIB_heure_maghrib.setText(gb_heure_maghrib)
                        if i == 6:
                            gb_heure_ishae = x.getText()
                            self.LIB_heure_ishae.setText(gb_heure_ishae)

            except Exception as e:
                print("Exception: ", e)

        if mode =="Offline":
            # get from database
            print("Mode Offline")
            # definir la période en se basant sur l'heure
            heure_actuel = datetime.datetime.now().strftime("%H:%M:%S")
            bheure_actuel = parser.parse(heure_actuel)
            lheure_actuel = bheure_actuel.strftime("%H:%M:%S")
            # print("heure actuel: ", lheure_actuel)
            if self.debug: print("heure_actuel>>", lheure_actuel)
            # kel journée
            kel_Date = datetime.datetime.now().strftime("%Y-%m-%d")
            # Récupérer les horaires tel date
            new_gb_heure_fajr = get_Horaire(kel_Date, "heure_fajr")
            gb_heure_fajr = new_gb_heure_fajr
            c = 0
            if new_gb_heure_fajr is None:
                print("Non Trouvé")
                bgo = False
                c = 0
            else:
                bgo = True
                self.LIB_heure_fajr.setText(gb_heure_fajr[:5])
                c = 1

            gb_heure_shourouq = get_Horaire(kel_Date, "heure_shourouq")
            if gb_heure_shourouq is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1

            gb_heure_dhuhr = get_Horaire(kel_Date, "heure_dhuhr")
            if gb_heure_dhuhr is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                self.LIB_heure_dhuhr.setText(gb_heure_dhuhr[:5])
                c = c + 1

            gb_heure_asr = get_Horaire(kel_Date, "heure_asr")
            if gb_heure_asr is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                self.LIB_heure_asr.setText(gb_heure_asr[:5])
                c = c + 1

            gb_heure_maghrib = get_Horaire(kel_Date, "heure_maghrib")
            if gb_heure_maghrib is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1
                self.LIB_heure_maghrib.setText(gb_heure_maghrib[:5])

            gb_heure_ishae = get_Horaire(kel_Date, "heure_ishae")
            if gb_heure_ishae is None:
                if self.debug: print("Non Trouvé")
                bgo = False
                c = c - 1
            else:
                bgo = True
                c = c + 1
                self.LIB_heure_ishae.setText(gb_heure_ishae[:5])

            if bgo:
                if debug: print("c: ", c)
                if c != 6:
                    bgo = False

            if bgo:
                # calcul difference
                if new_gb_heure_fajr < heure_actuel <= gb_heure_dhuhr:
                    self.LIB_DT_2.setText("Dhuhr")
                    if gb_heure_dhuhr == heure_actuel:
                        if self.debug: print("c'est l'heure dhuhr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_dhuhr = parser.parse(gb_heure_dhuhr)
                        lheure_dhuhr = bheure_dhuhr.strftime("%H:%M:%S")
                        print("Heure dhuhr: ", lheure_dhuhr)
                        time_diff = (parser.parse(lheure_dhuhr) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                # tester si l'heure actuel < heure asr
                if gb_heure_dhuhr < heure_actuel <= gb_heure_asr:
                    print("l'heure asr")
                    self.LIB_DT_2.setText("Asr")
                    if gb_heure_asr == heure_actuel:
                        print("c'est l'heure asr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        # calcul difference to asr
                        bheure_asr = parser.parse(gb_heure_asr)
                        lheure_asr = bheure_asr.strftime("%H:%M:%S")
                        print("Heure asr: ", lheure_asr)
                        time_diff = (parser.parse(lheure_asr) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_asr < heure_actuel <= gb_heure_maghrib:
                    # heure el maghrib
                    self.LIB_DT_2.setText("Maghrib")
                    if gb_heure_maghrib == heure_actuel:
                        if self.debug: print("c'est l'heure maghrib")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_maghrib = parser.parse(gb_heure_maghrib)
                        lheure_maghrib = bheure_maghrib.strftime("%H:%M:%S")
                        print("Heure maghrib: ", lheure_maghrib)
                        # time_diff = lheure_actuel - lheure_maghrib
                        time_diff = (parser.parse(lheure_maghrib) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_maghrib < heure_actuel <= gb_heure_ishae:

                    self.LIB_DT_2.setText("Isha")

                    if gb_heure_ishae == heure_actuel:
                        if self.debug: print("c'est l'heure ishae")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        bheure_ishae = parser.parse(gb_heure_ishae)
                        lheure_ishae = bheure_ishae.strftime("%H:%M:%S")
                        print("Heure ishae: ", lheure_ishae)
                        time_diff = (parser.parse(lheure_ishae) - parser.parse(lheure_actuel))
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

                if gb_heure_ishae < heure_actuel:

                    # Récupération Fajr next day

                    curr_date = datetime.datetime.now().strftime("%Y-%m-%d")
                    today = parser.parse(curr_date + " " + heure_actuel)
                    curr_date_temp = datetime.datetime.strptime(curr_date, "%Y-%m-%d")
                    new_date = curr_date_temp + datetime.timedelta(days=1)
                    new_date = new_date.strftime("%Y-%m-%d")
                    if self.debug: print("new date: ", new_date)
                    new_gb_heure_fajr = get_Horaire(new_date, "heure_fajr")
                    self.LIB_heure_fajr.setText(new_gb_heure_fajr[:5])
                    new_date_heure_fajr = new_date + " " + new_gb_heure_fajr
                    bday = parser.parse(new_date_heure_fajr)
                    if self.debug: print("New time: ", new_date_heure_fajr)
                    self.LIB_DT_2.setText("Fajr")
                    if new_gb_heure_fajr == heure_actuel:
                        if self.debug: print("c'est l'heure el fajr")
                        self.LIB_DT_4.setText("حان وقت الاذان")
                        self.playAudioFile()
                    else:
                        # calcul difference en heure entre les deux date
                        time_diff = bday - today
                        tsecs = time_diff.total_seconds()
                        tmins = tsecs / 60
                        thrs = tsecs / (60 * 60)
                        res = datetime.timedelta(seconds=tsecs)
                        print("Temps_Restant>> ", res)
                        Temps_Restant = str(res).zfill(8)
                        self.LIB_remining_Time.setText(Temps_Restant)

if __name__ == "__main__":
    config = configparser.ConfigParser()
    file = pathlib.Path("config/config.ini")
    if not file.exists():
        print("Configuration File not exist")
        # Create the file
        config['DEFAULT'] = {'version': '23.1.09.0',
                             'Url_WS1': 'https://timesprayer.com/en/prayer-times-in-',
                             'Url_WS2': 'https://www.prayertimes.org/morocco',
                             'timeout': '1',
                             'city': 'casablanca',
                             'background_image_url': 'jpg/DSC_0001.jpg',
                             'azan_mp3': 'mp3/azan1.mp3',
                             'fajr_image': 'svg/fajr.svg',
                             'dhuhr_image': 'svg/dhuhr.svg',
                             'asr_image': 'svg/asr.svg',
                             'maghrib_image': 'svg/maghrib.svg',
                             'isha_image': 'svg/isha.svg',
                             'remining_Time': '00:00:00',
                             'ResolutionW': '1266',
                             'ResolutionH': '768',
                             'mode': 'Online',
                             'note_Journee': '',
                             'note_Externe': '',
                             }

        with open('config/config.ini', 'w') as configfile:
            config.write(configfile)
        os.execv(sys.executable, [sys.executable] + sys.argv)

    else:
        config.read('config/config.ini')
        MONVER = config['DEFAULT']['version']
        MONURL1 = config['DEFAULT']['Url_WS1']
        MONURL2 = config['DEFAULT']['Url_WS2']
        time_out = config['DEFAULT']['timeout']
        city = config['DEFAULT']['city']
        ResolutionW = config['DEFAULT']['ResolutionW']
        ResolutionH = config['DEFAULT']['ResolutionH']
        background_image_url = config['DEFAULT']['background_image_url']
        azan_mp3 = config['DEFAULT']['azan_mp3']
        fajr_image = config['DEFAULT']['fajr_image']
        dhuhr_image = config['DEFAULT']['dhuhr_image']
        asr_image = config['DEFAULT']['asr_image']
        maghrib_image = config['DEFAULT']['maghrib_image']
        isha_image = config['DEFAULT']['isha_image']
        remining_Time = config['DEFAULT']['remining_Time']
        note_Journee = config['DEFAULT']['note_Journee']
        note_Externe = config['DEFAULT']['note_Externe']
        mode = config['DEFAULT']['mode']
    if not createConnection():
        print("Unable to connect to the database")
        sys.exit(1)
    else:
        # Create table horaire
        createTableQuery = QSqlQuery()
        createTableQuery.exec(
            """
            CREATE TABLE IF NOT EXISTS Horaires (
                id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
                Jour_Date DATE NOT NULL,
                Jour  VARCHAR(40) NOT NULL,
                Mois_AR VARCHAR(40) NOT NULL,
                Mois_FR VARCHAR(40) NOT NULL,
                heure_fajr TIME NOT NULL,
                heure_shourouq TIME,
                heure_dhuhr TIME NOT NULL,
                heure_asr TIME NOT NULL,
                heure_maghrib TIME NOT NULL,
                heure_ishae TIME NOT NULL
            )
            """
        )
        createTableQuery.clear()
    # Mise à jour des Paramètres, création si non existant
    if Insert_Update_Param():
        MONVER = get_param("version")
        MONURL1 = get_param("Url_WS1")
        MONURL2 = get_param("Url_WS2")
        time_out = get_param("timeout")
        city = get_param("city")
        ResolutionW = get_param("ResolutionW")
        ResolutionH = get_param("ResolutionH")
        background_image_url = get_param("background_image_url")
        azan_mp3 = get_param("azan_mp3")
        remining_Time = get_param("remining_Time")
        note_Journee = get_param("note_Journee")
        note_Externe = get_param("note_Externe")
        mode = get_param("mode")
        # Close Connection
        # The connection is closed but still in the list of connections
        # print(QSqlDatabase.connectionNames())
        # Remove the default connection
        QSqlDatabase.removeDatabase(QSqlDatabase.database().connectionName())
        # The connection is no longer in the list of connections
        # print(QSqlDatabase.connectionNames()
        # remining_Time = "00:05:00"
    app = QtWidgets.QApplication(sys.argv)

    w = MainWindow()
    w.show()
    if not check_if_existe(azan_mp3):
        sys.exit(1)
    if not check_if_existe(background_image_url):
        sys.exit(1)
    sys.exit(app.exec_())
